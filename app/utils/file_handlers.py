import os
import shutil
from typing import Optional
from fastapi import UploadFile
import PyPDF2
import docx
import pandas as pd
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)

class FileHandler:
    def __init__(self, upload_dir: str):
        self.upload_dir = upload_dir
        
    async def save_file(self, file: UploadFile, filename: str) -> str:
        """Save uploaded file to disk"""
        file_path = os.path.join(self.upload_dir, filename)
        
        # Ensure unique filename
        counter = 1
        while os.path.exists(file_path):
            name, ext = os.path.splitext(filename)
            file_path = os.path.join(self.upload_dir, f"{name}_{counter}{ext}")
            counter += 1
            
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        return file_path
    
    async def extract_text(self, file_path: str) -> str:
        """Extract text from various file formats"""
        ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if ext == '.pdf':
                return self._extract_text_from_pdf(file_path)
            elif ext in ['.docx', '.doc']:
                return self._extract_text_from_docx(file_path)
            elif ext in ['.xlsx', '.xls']:
                return self._extract_text_from_excel(file_path)
            elif ext == '.csv':
                return self._extract_text_from_csv(file_path)
            elif ext in ['.txt', '.md']:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            else:
                return ""  # For images, return empty string
        except Exception as e:
            logger.error(f"Text extraction failed for {file_path}: {str(e)}")
            return ""
    
    def _extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF using PyPDF2"""
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    
    def _extract_text_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        doc = docx.Document(file_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    
    def _extract_text_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        text_parts = []
        
        # Read all sheets
        try:
            df_dict = pd.read_excel(file_path, sheet_name=None)
            for sheet_name, df in df_dict.items():
                text_parts.append(f"Sheet: {sheet_name}")
                text_parts.append(df.to_string())
        except:
            # Fallback for older Excel formats
            wb = load_workbook(file_path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                text_parts.append(f"Sheet: {sheet}")
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text_parts.append("\t".join([str(cell) if cell else "" for cell in row]))
                    
        return "\n".join(text_parts)
    
    def _extract_text_from_csv(self, file_path: str) -> str:
        """Extract text from CSV file"""
        df = pd.read_csv(file_path)
        return df.to_string()